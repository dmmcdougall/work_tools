"""
This script collects the head timesheets at puts them into a single
.xlsx for editing.
"""

# imported from standard library
import datetime as dt
import openpyxl
import os
from shutil import copyfile
from tqdm import trange, tqdm
from win32com.client import Dispatch
import xlsxwriter

# imported from third party repos

# imported from local directories
import config as cfg

def main():
    # get the wk ending info from the usr
    my_date = input("""Give me the date string you would like to use as the week ending date.
The excepted format is a 3 character month name, a day number, a comma, and a year. 
An example of the excepted format is...

Sep 14, 1985

Please enter your week ending date. 
""")
    print()

    my_date_comma = my_date.replace(',', '')
    wk_end = dt.datetime.strptime(my_date, "%b %d, %Y")
    wk_begin = wk_end - dt.timedelta(days=6)

    # create dirs and files
    dir_name = f"{cfg.desktop_dir}\\Week Ends {my_date}"
    os.mkdir(dir_name)

    read_dir = cfg.my_dir 
    read_list = os.listdir(read_dir)
    num_sheets = len(read_list)

    user_y_n = input("""Would you like me to grab files from the current_timesheet folder?
If not, you will need to move them into the created directory yourself.  

Enter Y or N now.
""")
    affirmitives = ['yes','Yes', 'y', 'Y']
    if user_y_n in affirmitives:
        print()
        print('Copying the files to the new directory...')
        for my_file in tqdm(read_list):
            file_path = read_dir + '\\' + my_file
            save_path = dir_name + '\\' + my_file
            copyfile(file_path, save_path)
    else:
        print("Okay I'll wait.  Press ENTER when done")
        input() 
        
    
    filename = f"House Crew Timesheets - Week Ends {my_date_comma}.xlsx"
    workbook = xlsxwriter.Workbook(dir_name+"\\"+filename)
    write_file = dir_name +"\\"+filename
    workbook.close()

    print("Done setup, let's do some work...")

    # take each timesheet, place it in the write file
    xl = Dispatch("Excel.Application")
    print()
    print('Placing the single timesheets into the master workbook...')
    for i in trange(num_sheets):
        read_file = (read_dir + '\\' + read_list[i])
        wb1 = xl.Workbooks.Open(Filename=read_file)
        wb2 = xl.Workbooks.Open(Filename=write_file)
        ws1 = wb1.Worksheets(1)
        ws1.Copy(Before=wb2.Worksheets(i+1))
        wb2.Close(SaveChanges=True)
    xl.Quit()

    # fix the dates in the timesheets
    print()
    print('Setting the date on the sheets...')
    for i in trange(num_sheets):
        wb = openpyxl.load_workbook(write_file)
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[i]]
        Sheet1 .cell(row = 21, column = 1).value = wk_begin
        wb.save(write_file) 

    # remove the 'Sheet1' from intialization process
    wb = openpyxl.load_workbook(write_file)
    sheet_list = wb.sheetnames
    wb.remove(wb['Sheet1'])
    wb.save(write_file) 

if __name__ == '__main__':
    print()
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print("                      file launched")
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print()
    main()
    print()
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print("                        VICTORY!")
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print()